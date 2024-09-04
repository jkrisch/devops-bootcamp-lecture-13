import openpyxl as op

workbook = op.load_workbook("employees.xlsx")
employees = workbook["Sheet1"]

employees_list = []
for row in range(2,employees.max_row):
    current_employee = {}
    current_employee["name"] = employees.cell(row=row, column=1).value
    current_employee["xp"] = int(employees.cell(row=row, column=2).value)
    employees_list.append(current_employee)

employee_list_sorted = sorted(employees_list, key=lambda d: d['xp'], reverse=True)

print(employee_list_sorted)

new_wb = op.Workbook()

ws = new_wb.active

ws.append(["Name", "Years of Experience"])

for employee in employee_list_sorted:
    ws.append([employee["name"],employee["xp"]])

new_wb.save("employees_sorted.xlsx")